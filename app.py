import io
import re
from pathlib import Path

import pdfplumber
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 30 * 1024 * 1024
CORS(app)

HEADING_FILL = PatternFill(start_color='FF2563EB', end_color='FF2563EB', fill_type='solid')
HEADING_FONT = Font(bold=True, color='FFFFFFFF')
THIN = Side(style='thin', color='FFD1D5DB')
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top', horizontal='left')
WRAP_CENTER = Alignment(wrap_text=True, vertical='center', horizontal='center')
NUMERIC_RE = re.compile(r'^-?\d[\d,]*\.?\d*$')
INDEX_RE = re.compile(r'^\(\s*\d+\s*\)$')

# The source PDF uses a real ruled/grid table. Using pdfplumber's line-based
# table finder is much safer than reconstructing columns from text positions.
TABLE_SETTINGS = {
    'vertical_strategy': 'lines',
    'horizontal_strategy': 'lines',
    'intersection_tolerance': 5,
    'snap_tolerance': 3,
    'join_tolerance': 3,
    'edge_min_length': 20,
}


def clean_cell(value):
    if value is None:
        return ''
    # Preserve intentional newlines inside a PDF cell, but normalize spaces.
    lines = []
    for line in str(value).replace('\r', '\n').split('\n'):
        line = re.sub(r'[ \t]+', ' ', line).strip()
        if line:
            lines.append(line)
    return '\n'.join(lines)


def is_english_text(text):
    s = re.sub(r'\s+', '', text or '')
    if not s:
        return False
    ascii_chars = sum(1 for ch in s if ord(ch) < 128)
    return ascii_chars / len(s) >= 0.60


def is_index_row(row, expected_cols=None):
    if not row:
        return False
    cells = [clean_cell(x) for x in row]
    nonblank = [x for x in cells if x]
    if expected_cols and len(cells) != expected_cols:
        return False
    return bool(nonblank) and all(INDEX_RE.fullmatch(x) for x in nonblank)


def is_real_english_header(rows):
    """A real Gazette header is followed by the (1)..(6) index row."""
    if len(rows) < 2:
        return False
    header = [clean_cell(x) for x in rows[0]]
    index = [clean_cell(x) for x in rows[1]]
    if len(header) != len(index):
        return False
    if not is_index_row(index, len(header)):
        return False
    return is_english_text(' '.join(header))


def normalize_rows(rows):
    if not rows:
        return []
    ncols = max(len(r) for r in rows)
    out = []
    for r in rows:
        vals = [clean_cell(x) for x in r]
        vals += [''] * (ncols - len(vals))
        out.append(vals[:ncols])
    return out


def numeric_or_price(text):
    return bool(NUMERIC_RE.fullmatch((text or '').strip()))


def break_after_comma(text):
    s = clean_cell(text)
    if not s or numeric_or_price(s):
        return s
    # Do not split decimal/thousands numeric fragments embedded in otherwise
    # textual cells; only normal comma separators are converted.
    return ',\n'.join(part.strip() for part in s.split(',') if part.strip())


def table_geometry(table):
    x0, top, x1, bottom = table.bbox
    width = max(1.0, x1 - x0)
    return x0, x1, width


def same_table_geometry(a, b, tolerance=0.08):
    """True when two ruled tables occupy essentially the same six-column grid."""
    ax0, ax1, aw = table_geometry(a)
    bx0, bx1, bw = table_geometry(b)
    return (
        abs(ax0 - bx0) <= max(8, aw * tolerance)
        and abs(ax1 - bx1) <= max(8, aw * tolerance)
        and abs(aw - bw) <= max(12, aw * tolerance)
    )


def choose_tables_on_page(page):
    """Return ruled tables only. This excludes ordinary page text/footnotes."""
    return page.find_tables(table_settings=TABLE_SETTINGS)


def choose_english_header_table(tables):
    for table in tables:
        rows = normalize_rows(table.extract())
        if is_real_english_header(rows):
            return table, rows
    return None, None


def merge_continuation(previous, row):
    """Merge a row that starts with a blank Sl. No. into the previous row."""
    if not row or not previous:
        return False
    if row[0].strip() != '':
        return False
    if not any(c.strip() for c in row[1:]):
        return False

    for i, value in enumerate(row):
        value = clean_cell(value)
        if not value:
            continue
        if previous[i]:
            previous[i] = previous[i].rstrip() + '\n' + value
        else:
            previous[i] = value
    return True


def is_new_data_row(row):
    return bool(row and any(c.strip() for c in row))


def extract_english_tables(pdf_stream, english_only=True, merge_continuations=True):
    """
    Extract only ruled tables whose first page contains an English header +
    (1)..(n) index row. Subsequent pages are accepted only when their ruled
    table has the same geometry/column count. Because extraction is limited
    to the table's bounding box, page numbers, running headers and notes
    outside the grid are never included.
    """
    logical_tables = []
    current = None
    current_header = None
    current_geometry = None
    current_ncols = None

    with pdfplumber.open(pdf_stream) as pdf:
        for page_no, page in enumerate(pdf.pages, start=1):
            tables = choose_tables_on_page(page)
            if not tables:
                continue

            # A printed header/index row identifies the beginning of a logical
            # table. There may be Hindi tables earlier in the same PDF; those
            # are ignored when english_only=True.
            header_table, header_rows = choose_english_header_table(tables)

            if header_table is not None:
                if current is not None:
                    logical_tables.append(current)
                header = header_rows[0]
                data = header_rows[2:]
                current = [header]
                current_header = header
                current_geometry = header_table
                current_ncols = len(header)
                for row in data:
                    if is_new_data_row(row):
                        current.append(row)
                continue

            # No header on this page: only continue an already-open English
            # table if the ruled grid matches the same column structure.
            if current is None or not english_only:
                continue

            candidate = None
            candidate_rows = None
            for table in tables:
                rows = normalize_rows(table.extract())
                if len(rows) == 0 or len(rows[0]) != current_ncols:
                    continue
                if same_table_geometry(current_geometry, table):
                    candidate = table
                    candidate_rows = rows
                    break

            if candidate is None:
                # Do not attach unrelated tables, notes, or page text.
                continue

            for row in candidate_rows:
                if not is_new_data_row(row):
                    continue
                if merge_continuations and len(current) > 1 and merge_continuation(current[-1], row):
                    continue
                current.append(row)

    if current is not None:
        logical_tables.append(current)

    # A final guard: only keep tables with an English header and at least one
    # data row. This also prevents any accidental non-table content from being
    # written to Excel.
    if english_only:
        logical_tables = [t for t in logical_tables if len(t) > 1 and is_english_text(' '.join(t[0]))]
    return logical_tables


def write_workbook(tables, comma_breaks=True):
    wb = Workbook()
    wb.remove(wb.active)

    for idx, table in enumerate(tables, start=1):
        ws = wb.create_sheet((f'Table {idx}' if len(tables) > 1 else 'Table')[:31])
        header = table[0]
        ncols = len(header)

        ws.append(header)
        for c in range(1, ncols + 1):
            cell = ws.cell(1, c)
            cell.fill = HEADING_FILL
            cell.font = HEADING_FONT
            cell.alignment = WRAP_CENTER
            cell.border = CELL_BORDER

        for row in table[1:]:
            values = [break_after_comma(c) if comma_breaks else clean_cell(c) for c in row]
            ws.append(values)
            r = ws.max_row
            for c in range(1, ncols + 1):
                cell = ws.cell(r, c)
                cell.alignment = WRAP
                cell.border = CELL_BORDER

        for c in range(1, ncols + 1):
            letter = get_column_letter(c)
            max_len = 10
            for r in range(1, ws.max_row + 1):
                value = ws.cell(r, c).value or ''
                for line in str(value).split('\n'):
                    max_len = max(max_len, len(line))
            ws.column_dimensions[letter].width = min(max_len + 2, 45)

        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = ws.dimensions
        ws.row_dimensions[1].height = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@app.get('/health')
def health():
    return jsonify({'status': 'ok', 'engine': 'pdfplumber-line-table+openpyxl', 'version': '4'})


@app.post('/convert')
def convert():
    uploaded = request.files.get('file')
    if not uploaded or not uploaded.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Please upload a PDF file.'}), 400

    try:
        data = uploaded.read()
        if not data:
            return jsonify({'error': 'The uploaded PDF is empty.'}), 400

        english_only = request.form.get('english_only', '1') == '1'
        merge_continuations = request.form.get('merge_continuations', '1') == '1'
        comma_breaks = request.form.get('comma_breaks', '1') == '1'

        tables = extract_english_tables(
            io.BytesIO(data),
            english_only=english_only,
            merge_continuations=merge_continuations,
        )
        if not tables:
            return jsonify({'error': 'No ruled English table was found. Scanned/image-only or borderless tables require a different extraction method.'}), 422

        output = write_workbook(tables, comma_breaks=comma_breaks)
        filename = Path(uploaded.filename).stem + '_converted.xlsx'
        return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as exc:
        app.logger.exception('Conversion error')
        return jsonify({'error': f'Conversion failed: {exc}'}), 500


@app.errorhandler(413)
def too_large(_):
    return jsonify({'error': 'PDF is too large. Maximum size is 30 MB.'}), 413


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8000)
